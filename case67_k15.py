# -*- coding: utf-8 -*-
import os
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="temp內xlsx標準資料表整併(工作表名可選)", layout="wide")
st.title("📦 temp內所有.xlsx 標準資料表依欄位自動整併（v0_3_3）")

# =========================================================
# 路徑
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, "temp")
OUT_DIR = os.path.join(BASE_DIR, "out")

os.makedirs(TEMP_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

# =========================================================
# 工具函式
# =========================================================
def normalize_columns(cols):
    out = []
    for c in cols:
        s = "" if pd.isna(c) else str(c)
        s = s.replace("\n", " ").replace("\r", " ").strip()
        s = " ".join(s.split())
        out.append(s)
    return out


def is_standard_table(df: pd.DataFrame) -> bool:
    if df is None or not isinstance(df, pd.DataFrame):
        return False
    if df.shape[1] == 0:
        return False
    cols = normalize_columns(df.columns)
    if all(c == "" for c in cols):
        return False
    if df.shape[0] == 0:
        return False
    return True


def get_signature(columns, ignore_order=False):
    cols = normalize_columns(columns)
    return tuple(sorted(cols)) if ignore_order else tuple(cols)


def list_xlsx_files(folder):
    return sorted([
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ])


def build_source_value(file_path, sheet_name):
    return f"{file_path}_{sheet_name}"


def add_source_id(df):
    """
    來源id為中繼欄位，只用於著色。
    同一個結果df內，相同來源會得到相同id。
    """
    df = df.copy()
    if "來源" not in df.columns:
        df["來源id"] = None
        return df

    source_map = {}
    next_id = 1
    source_ids = []

    for val in df["來源"].astype(str):
        if val not in source_map:
            source_map[val] = next_id
            next_id += 1
        source_ids.append(source_map[val])

    df["來源id"] = source_ids
    return df


def get_fill_by_source_id(source_id):
    color_list = [
        "FFF2CC",  # 淡黃
        "E2F0D9",  # 淡綠
        "DDEBF7",  # 淡藍
        "FCE4D6",  # 淡橘
        "E4DFEC",  # 淡紫
        "D9EAD3",  # 綠灰
        "F4CCCC",  # 淡紅
        "D0E0E3",  # 淡青灰
        "FCE5CD",  # 淡杏
        "EAD1DC",  # 淡粉紫
    ]

    if source_id is None:
        return PatternFill(fill_type="solid", fgColor="FFFFFF")

    try:
        idx = (int(source_id) - 1) % len(color_list)
    except Exception:
        idx = 0

    return PatternFill(fill_type="solid", fgColor=color_list[idx])


def style_ws(ws, export_df, source_ids):
    """
    1. A2凍結
    2. 標題列淺藍
    3. 資料列依來源id著色
    4. 匯出時不包含來源id欄位
    """
    ws.freeze_panes = "A2"

    fill_header = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = bold_font

    for row_idx, sid in enumerate(source_ids, start=2):
        fill = get_fill_by_source_id(sid)
        for col_idx in range(1, len(export_df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    for i, col in enumerate(export_df.columns, start=1):
        max_len = len(str(col))
        if len(export_df) > 0:
            sample = export_df[col].astype(str).head(300)
            if len(sample) > 0:
                max_len = max(max_len, sample.map(len).max())
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)


def sanitize_sheet_name(name: str) -> str:
    """
    Excel 工作表名稱限制：
    - 不可含 \ / ? * [ ] :
    - 最長31字
    - 不可空白
    """
    if name is None:
        name = ""
    name = str(name).strip()
    name = re.sub(r'[\\\/\?\*\[\]\:]', "_", name)
    name = name.replace("\n", " ").replace("\r", " ").strip()
    name = " ".join(name.split())
    if name == "":
        name = "Sheet"
    return name[:31]


def make_unique_sheet_name(base_name: str, used_names: set) -> str:
    """
    若重複，自動補 (2), (3) ...
    並控制總長 <= 31
    """
    base_name = sanitize_sheet_name(base_name)
    if base_name not in used_names:
        used_names.add(base_name)
        return base_name

    n = 2
    while True:
        suffix = f"({n})"
        allow_len = 31 - len(suffix)
        candidate = f"{base_name[:allow_len]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        n += 1


# =========================================================
# 核心：收集標準資料表
# =========================================================
def collect_standard_tables(temp_dir):
    table_objects = []
    preview_rows = []

    files = list_xlsx_files(temp_dir)

    for file_path in files:
        file_name = os.path.basename(file_path)

        try:
            xls = pd.ExcelFile(file_path)
        except Exception as e:
            st.warning(f"讀取檔案失敗：{file_name}｜{e}")
            continue

        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            except Exception as e:
                st.warning(f"讀取工作表失敗：{file_name}｜{sheet_name}｜{e}")
                continue

            if not is_standard_table(df):
                continue

            df = df.copy()
            df.columns = normalize_columns(df.columns)
            df["來源"] = build_source_value(file_path, sheet_name)

            table_objects.append({
                "df": df,
                "file_path": file_path,
                "file_name": file_name,
                "sheet_name": sheet_name,
            })

            preview_rows.append({
                "來源檔案": file_name,
                "來源工作表": sheet_name,
                "資料列數": len(df),
                "欄位數": len(df.columns),
                "欄位名稱": " | ".join(df.columns),
            })

    preview_df = pd.DataFrame(preview_rows)
    return table_objects, preview_df


# =========================================================
# 核心：依欄位整併 + 工作表命名模式
# =========================================================
def merge_tables_by_columns(table_objects, ignore_order=False, sheet_name_mode="df序號"):
    """
    sheet_name_mode:
    - df序號
    - first工作表名
    """
    title_bank = {}         # signature -> result_sheet_name
    result_dfs = {}         # result_sheet_name -> dataframe
    first_sheet_name_map = {}   # signature -> first original sheet_name
    used_sheet_names = set()
    summary_rows = []

    for item in table_objects:
        df = item["df"]
        orig_sheet_name = item["sheet_name"]

        sig = get_signature(df.columns, ignore_order=ignore_order)

        if sig in title_bank:
            key = title_bank[sig]
            canonical_cols = list(result_dfs[key].columns)
            df_aligned = df.reindex(columns=canonical_cols)
            result_dfs[key] = pd.concat(
                [result_dfs[key], df_aligned],
                ignore_index=True,
                sort=False
            )
        else:
            first_sheet_name_map[sig] = orig_sheet_name

            if sheet_name_mode == "first工作表名":
                new_key = make_unique_sheet_name(orig_sheet_name, used_sheet_names)
            else:
                new_key = make_unique_sheet_name(f"df{len(title_bank) + 1}", used_sheet_names)

            title_bank[sig] = new_key
            result_dfs[new_key] = df.copy()

    # 為每個結果df補上來源id（中繼欄位）
    for key in list(result_dfs.keys()):
        result_dfs[key] = add_source_id(result_dfs[key])

    for sig, key in title_bank.items():
        df = result_dfs[key]
        display_cols = [c for c in df.columns if c != "來源id"]
        summary_rows.append({
            "結果工作表": key,
            "首個來源工作表名": first_sheet_name_map.get(sig, ""),
            "資料列數": len(df),
            "欄位數": len(display_cols),
            "欄位名稱": " | ".join([str(c) for c in display_cols]),
        })

    summary_df = pd.DataFrame(summary_rows)
    return result_dfs, summary_df


def export_excel(result_dfs, out_path):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in result_dfs.items():
            export_df = df.copy()
            source_ids = export_df["來源id"].tolist() if "來源id" in export_df.columns else [None] * len(export_df)

            if "來源id" in export_df.columns:
                export_df = export_df.drop(columns=["來源id"])

            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]
            style_ws(ws, export_df, source_ids)


# =========================================================
# Sidebar
# =========================================================
st.sidebar.header("⚙️ 控制區")
st.sidebar.write(f"TEMP：`{TEMP_DIR}`")
st.sidebar.write(f"OUT：`{OUT_DIR}`")

ignore_order = st.sidebar.checkbox("忽略欄位順序（只比對欄位名稱）", value=False)

sheet_name_mode_label = st.sidebar.selectbox(
    "結果工作表命名方式",
    ["df1~dfn", "first工作表名"],
    index=0
)

sheet_name_mode = "df序號" if sheet_name_mode_label == "df1~dfn" else "first工作表名"

run_btn = st.sidebar.button("▶️ 開始演算", use_container_width=True)
export_btn = st.sidebar.button("💾 匯出結果", use_container_width=True)

# =========================================================
# Session State
# =========================================================
if "case67_v0_3_3_result_dfs" not in st.session_state:
    st.session_state["case67_v0_3_3_result_dfs"] = {}

if "case67_v0_3_3_summary_df" not in st.session_state:
    st.session_state["case67_v0_3_3_summary_df"] = pd.DataFrame()

if "case67_v0_3_3_preview_df" not in st.session_state:
    st.session_state["case67_v0_3_3_preview_df"] = pd.DataFrame()

# =========================================================
# 執行
# =========================================================
if run_btn:
    tables, preview_df = collect_standard_tables(TEMP_DIR)
    result_dfs, summary_df = merge_tables_by_columns(
        tables,
        ignore_order=ignore_order,
        sheet_name_mode=sheet_name_mode
    )

    st.session_state["case67_v0_3_3_result_dfs"] = result_dfs
    st.session_state["case67_v0_3_3_summary_df"] = summary_df
    st.session_state["case67_v0_3_3_preview_df"] = preview_df

    if len(result_dfs) == 0:
        st.warning("temp 內沒有可整併的標準資料表。")
    else:
        st.success(f"完成：整併為 {len(result_dfs)} 個結果工作表")

# =========================================================
# 顯示
# =========================================================
result_dfs = st.session_state["case67_v0_3_3_result_dfs"]
summary_df = st.session_state["case67_v0_3_3_summary_df"]
preview_df = st.session_state["case67_v0_3_3_preview_df"]

if len(result_dfs) == 0:
    st.info("請先將所有 .xlsx 放入 temp 資料夾，再按「開始演算」。")
else:
    c1, c2, c3 = st.columns(3)
    c1.metric("標準資料表數", len(preview_df))
    c2.metric("整併後工作表數", len(result_dfs))
    c3.metric("總資料列數", sum(len(df) for df in result_dfs.values()))

    st.subheader("① 原始資料表盤點")
    st.dataframe(preview_df, use_container_width=True, height=220)

    st.subheader("② 整併摘要")
    st.dataframe(summary_df, use_container_width=True, height=220)

    st.subheader("③ 整併結果預覽")
    tabs = st.tabs(list(result_dfs.keys()))
    for tab, name in zip(tabs, result_dfs.keys()):
        with tab:
            df_show = result_dfs[name].copy()
            if "來源id" in df_show.columns:
                df_show = df_show.drop(columns=["來源id"])
            st.write(f"**{name}**｜rows={len(df_show)}｜cols={len(df_show.columns)}")
            st.dataframe(df_show, use_container_width=True, height=500)

# =========================================================
# 匯出
# =========================================================
if export_btn:
    result_dfs = st.session_state["case67_v0_3_3_result_dfs"]

    if len(result_dfs) == 0:
        st.warning("請先執行演算")
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(
            OUT_DIR,
            f"case67_v0_3_3_{ts}_整併結果.xlsx"
        )

        try:
            export_excel(result_dfs, out_path)
            st.success(f"匯出完成：{out_path}")
        except Exception as e:
            st.error(f"匯出失敗：{e}")